import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from tkinter import filedialog


class SupplierView(tk.Frame):

    def __init__(self, parent, db):
        super().__init__(parent, bg="#f9fafb")

        self.db = db
        self.selected_id = None

        self.build_ui()
        self.selected_suppliers = set()

    # ======================================
    # UI
    # ======================================

    def build_ui(self):

        header = tk.Frame(self, bg="#f9fafb")
        header.pack(fill="x")

        tk.Label(
            header, text="Master Pemasok", font=("Segoe UI", 20, "bold"), bg="#f9fafb"
        ).pack(side="left", padx=10, pady=10)

        btn_frame = tk.Frame(self, bg="#f9fafb")
        btn_frame.pack(fill="x", pady=5)

        tk.Button(
            btn_frame,
            text="Tambah Pemasok",
            bg="#10b981",
            width=16,
            command=self.popup_add,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Edit Pemasok",
            bg="#f59e0b",
            width=16,
            command=self.popup_edit,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Hapus Pemasok",
            bg="#ef4444",
            width=16,
            command=self.delete_supplier,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Cetak",
            bg="#8b5cf6",
            width=16,
            command=self.print_suppliers,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Import XLSX",
            bg="#6366f1",
            width=16,
            command=self.import_suppliers,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Refresh",
            bg="#3b82f6",
            width=16,
            command=self.load_data,
        ).pack(side="right", padx=5)

        # SEARCH
        search_frame = tk.Frame(self, bg="#f9fafb")
        search_frame.pack(fill="x", pady=5)

        self.select_all_var = tk.BooleanVar()

        def toggle_select_all():

            checked = self.select_all_var.get()

            self.selected_suppliers.clear()

            if checked:
                for item in self.tree.get_children():
                    vals = self.tree.item(item, "values")
                    code = vals[2]
                    self.selected_suppliers.add(code)

            self.refresh_checkboxes()

        tk.Checkbutton(
            search_frame,
            text="All",
            variable=self.select_all_var,
            command=toggle_select_all,
            bg="#f9fafb",
        ).pack(side="left", padx=(10, 5))

        tk.Label(search_frame, text="Cari pemasok:", bg="#f9fafb").pack(
            side="left", padx=(5, 2)
        )

        self.search_entry = tk.Entry(search_frame, width=30)
        self.search_entry.pack(side="left", padx=5)

        self.search_entry.bind("<KeyRelease>", lambda e: self.load_data())

        tk.Label(search_frame, text="Status:", bg="#f9fafb").pack(
            side="left", padx=(15, 2)
        )

        self.status_filter = ttk.Combobox(
            search_frame,
            values=["SEMUA", "AKTIF", "NONAKTIF"],
            width=12,
            state="readonly",
        )

        self.status_filter.set("AKTIF")
        self.status_filter.pack(side="left", padx=5)

        self.status_filter.bind("<<ComboboxSelected>>", lambda e: self.load_data())

        # TABLE
        self.tree = ttk.Treeview(
            self,
            columns=(
                "check",
                "no",
                "code",
                "name",
                "phone",
                "email",
                "npwp",
                "pic",
                "address",
                "status",
            ),
            show="headings",
            height=20,
        )

        headers = [
            ("check", "", 40),
            ("no", "No", 60),
            ("code", "Kode", 120),
            ("name", "Nama", 200),
            ("phone", "Telepon", 140),
            ("email", "Email", 200),
            ("npwp", "NPWP", 160),
            ("pic", "PIC", 160),
            ("address", "Alamat", 200),
            ("status", "Status", 100),
        ]

        for c, h, w in headers:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w)

        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        # style untuk supplier nonaktif
        self.tree.tag_configure("inactive", foreground="gray")

        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Button-1>", self.toggle_checkbox)

        self.load_data()

    # ======================================
    # LOAD DATA
    # ======================================

    def load_data(self):

        for i in self.tree.get_children():
            self.tree.delete(i)

        keyword = self.search_entry.get().lower()

        rows = self.db.execute("SELECT * FROM suppliers ORDER BY name").fetchall()

        status_filter = self.status_filter.get()

        no = 1

        for r in rows:

            if status_filter == "AKTIF" and r["is_active"] == 0:
                continue

            if status_filter == "NONAKTIF" and r["is_active"] == 1:
                continue

            name = (r["name"] or "").lower()

            if keyword and keyword not in name:
                continue

            is_active = r["is_active"]

            status = "Aktif" if is_active == 1 else "❌ Non Aktif"

            tag = ""
            name = r["name"]

            if is_active == 0:
                tag = "inactive"
                name = f"❌ {name}"

            self.tree.insert(
                "",
                "end",
                values=(
                    "☐",
                    no,
                    r["code"],
                    name,
                    r["phone"],
                    r["email"],
                    r["npwp"],
                    r["pic"],
                    r["address"],
                    status,
                ),
                tags=(tag,),
            )

            no += 1

    # ======================================
    # SELECT
    # ======================================

    def on_select(self, event=None):

        s = self.tree.selection()

        if not s:
            return

        vals = self.tree.item(s[0], "values")

        code = vals[2]

        row = self.db.execute(
            "SELECT id FROM suppliers WHERE code=?", (code,)
        ).fetchone()

        if row:
            self.selected_id = row["id"]

    # ======================================
    # ADD
    # ======================================

    def popup_add(self):

        win = tk.Toplevel(self)
        win.title("Tambah Pemasok")
        win.geometry("500x260")

        frame = tk.Frame(win)
        frame.pack(padx=20, pady=20)

        tk.Label(frame, text="Nama").grid(row=0, column=0)
        name = tk.Entry(frame, width=30)
        name.grid(row=0, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Telepon").grid(row=1, column=0)
        phone = tk.Entry(frame, width=30)
        phone.grid(row=1, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Email").grid(row=2, column=0)
        email = tk.Entry(frame, width=30)
        email.grid(row=2, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="NPWP").grid(row=3, column=0)
        npwp = tk.Entry(frame, width=30)
        npwp.grid(row=3, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="PIC").grid(row=0, column=2)
        pic = tk.Entry(frame, width=30)
        pic.grid(row=0, column=3, pady=5, padx=(0, 20))

        tk.Label(frame, text="Alamat").grid(row=1, column=2)
        addr = tk.Entry(frame, width=30)
        addr.grid(row=1, column=3, pady=5, padx=(0, 20))

        def save():

            row = self.db.execute("SELECT MAX(id) max_id FROM suppliers").fetchone()

            next_id = (row["max_id"] or 0) + 1

            code = f"SUP-{str(next_id).zfill(4)}"

            self.db.execute(
                """
                INSERT INTO suppliers
                (code,name,phone,email,npwp,pic,address)
                VALUES(?,?,?,?,?,?,?)
                """,
                (
                    code,
                    name.get(),
                    phone.get(),
                    email.get(),
                    npwp.get(),
                    pic.get(),
                    addr.get(),
                ),
                commit=True,
            )

            win.destroy()
            self.load_data()

        tk.Button(frame, text="Simpan", bg="#06b6d4", command=save).grid(
            row=5, column=3, pady=10
        )

    def popup_edit(self):

        if not self.selected_id:
            messagebox.showerror("Error", "Pilih pemasok dulu")
            return

        row = self.db.execute(
            "SELECT * FROM suppliers WHERE id=?", (self.selected_id,)
        ).fetchone()

        win = tk.Toplevel(self)
        win.title("Edit Pemasok")
        win.geometry("500x260")

        frame = tk.Frame(win)
        frame.pack(padx=20, pady=20)

        tk.Label(frame, text="Nama").grid(row=0, column=0)
        name = tk.Entry(frame, width=30)
        name.insert(0, row["name"] or "")
        name.grid(row=0, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Telepon").grid(row=1, column=0)
        phone = tk.Entry(frame, width=30)
        phone.insert(0, row["phone"] or "")
        phone.grid(row=1, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Email").grid(row=2, column=0)
        email = tk.Entry(frame, width=30)
        email.insert(0, row["email"] or "")
        email.grid(row=2, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="NPWP").grid(row=3, column=0)
        npwp = tk.Entry(frame, width=30)
        npwp.insert(0, row["npwp"] or "")
        npwp.grid(row=3, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="PIC").grid(row=0, column=2)
        pic = tk.Entry(frame, width=30)
        pic.insert(0, row["pic"] or "")
        pic.grid(row=0, column=3, pady=5, padx=(0, 20))

        tk.Label(frame, text="Alamat").grid(row=1, column=2)
        addr = tk.Entry(frame, width=30)
        addr.insert(0, row["address"] or "")
        addr.grid(row=1, column=3, pady=5, padx=(0, 20))

        inactive_var = tk.IntVar(value=0 if row["is_active"] == 1 else 1)

        tk.Checkbutton(frame, text="Nonaktifkan Pemasok", variable=inactive_var).grid(
            row=3, column=3, pady=5, padx=(0, 20)
        )

        def save():

            self.db.execute(
                """
                UPDATE suppliers
                SET name=?, phone=?, email=?, npwp=?, pic=?, address=?, is_active=?
                WHERE id=?
                """,
                (
                    name.get(),
                    phone.get(),
                    email.get(),
                    npwp.get(),
                    pic.get(),
                    addr.get(),
                    0 if inactive_var.get() else 1,
                    self.selected_id,
                ),
                commit=True,
            )

            messagebox.showinfo("Sukses", "Pemasok berhasil diupdate")

            win.destroy()
            self.load_data()

        tk.Button(frame, text="Update", bg="#06b6d4", command=save).grid(
            row=5, column=3, pady=10
        )

    # ======================================
    # DELETE
    # ======================================

    def delete_supplier(self):

        # jika ada checkbox dipilih
        if self.selected_suppliers:

            targets = list(self.selected_suppliers)

        else:

            # jika tidak ada checkbox, gunakan baris yang diseleksi
            sel = self.tree.selection()

            if not sel:
                messagebox.showerror("Error", "Pilih atau centang pemasok dulu")
                return

            vals = self.tree.item(sel[0], "values")
            targets = [vals[2]]  # kode supplier

        win = tk.Toplevel(self)
        win.title("Konfirmasi")
        win.geometry("350x150")
        win.configure(bg="#e5e7eb")

        tk.Label(
            win,
            text="Hapus Pemasok?",
            font=("Segoe UI", 14, "bold"),
            bg="#e5e7eb",
        ).pack(pady=20)

        btn_frame = tk.Frame(win, bg="#e5e7eb")
        btn_frame.pack()

        def confirm_delete():

            for code in targets:

                row = self.db.execute(
                    "SELECT id FROM suppliers WHERE code=?", (code,)
                ).fetchone()

                if row:

                    self.db.execute(
                        "DELETE FROM suppliers WHERE id=?",
                        (row["id"],),
                        commit=True,
                    )

            self.selected_suppliers.clear()

            win.destroy()

            messagebox.showinfo("Sukses", "Pemasok berhasil dihapus")

            self.load_data()

        tk.Button(
            btn_frame,
            text="Ya",
            width=10,
            bg="#16a34a",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=confirm_delete,
        ).pack(side="left", padx=10)

        tk.Button(
            btn_frame,
            text="Tidak",
            width=10,
            bg="#dc2626",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=win.destroy,
        ).pack(side="left", padx=10)

    # ======================================
    # CHECKBOX
    # ======================================

    def refresh_checkboxes(self):

        for item in self.tree.get_children():

            vals = list(self.tree.item(item, "values"))
            code = vals[2]

            if code in self.selected_suppliers:
                vals[0] = "☑"
            else:
                vals[0] = "☐"

            self.tree.item(item, values=vals)

    def toggle_checkbox(self, event):

        region = self.tree.identify("region", event.x, event.y)

        if region != "cell":
            return

        column = self.tree.identify_column(event.x)

        if column != "#1":
            return

        item = self.tree.identify_row(event.y)

        vals = list(self.tree.item(item, "values"))

        code = vals[2]

        if code in self.selected_suppliers:
            self.selected_suppliers.remove(code)
        else:
            self.selected_suppliers.add(code)

        self.refresh_checkboxes()

    def print_suppliers(self):

        win = tk.Toplevel(self)
        win.title("Konfirmasi")
        win.geometry("350x150")
        win.configure(bg="#e5e7eb")

        tk.Label(
            win,
            text="Cetak Data Pemasok?",
            font=("Segoe UI", 14, "bold"),
            bg="#e5e7eb",
        ).pack(pady=20)

        btn_frame = tk.Frame(win, bg="#e5e7eb")
        btn_frame.pack()

        def confirm_print():

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")],
                title="Simpan Data Pemasok",
            )

            if not filepath:
                return

            rows = self.db.execute(
                """
                SELECT code,name,phone,email,npwp,pic,address,is_active
                FROM suppliers
                ORDER BY name
                """
            ).fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = "Pemasok"

            headers = [
                "Kode",
                "Nama",
                "Telepon",
                "Email",
                "NPWP",
                "PIC",
                "Alamat",
                "Status",
            ]

            ws.append(headers)

            for r in rows:

                status = "AKTIF" if r["is_active"] == 1 else "NONAKTIF"

                ws.append(
                    [
                        r["code"],
                        r["name"],
                        r["phone"],
                        r["email"],
                        r["npwp"],
                        r["pic"],
                        r["address"],
                        status,
                    ]
                )

            wb.save(filepath)

            win.destroy()

            messagebox.showinfo("Sukses", "Data pemasok berhasil dicetak")

        tk.Button(
            btn_frame,
            text="Ya",
            width=10,
            bg="#16a34a",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=confirm_print,
        ).pack(side="left", padx=10)

        tk.Button(
            btn_frame,
            text="Tidak",
            width=10,
            bg="#dc2626",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=win.destroy,
        ).pack(side="left", padx=10)

    def import_suppliers(self):

        filepath = filedialog.askopenfilename(
            filetypes=[("Excel File", "*.xlsx")],
            title="Import Data Pemasok",
        )

        if not filepath:
            return

        wb = load_workbook(filepath)
        ws = wb.active

        insert_count = 0
        update_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):

            code = row[0]
            name = row[1]
            phone = row[2]
            email = row[3]
            npwp = row[4]
            pic = row[5]
            address = row[6]
            status = row[7] if len(row) > 7 else "AKTIF"

            is_active = 1
            if str(status).upper() == "NONAKTIF":
                is_active = 0

            if not code or not name:
                continue

            existing = self.db.execute(
                "SELECT id FROM suppliers WHERE code=?", (code,)
            ).fetchone()

            if existing:

                self.db.execute(
                    """
                    UPDATE suppliers
                    SET name=?,phone=?,email=?,npwp=?,pic=?,address=?,is_active=?
                    WHERE code=?
                    """,
                    (
                        name,
                        phone,
                        email,
                        npwp,
                        pic,
                        address,
                        is_active,
                        code,
                    ),
                    commit=True,
                )

                update_count += 1

            else:

                self.db.execute(
                    """
                    INSERT INTO suppliers
                    (code,name,phone,email,npwp,pic,address,is_active)
                    VALUES(?,?,?,?,?,?,?,?)
                    """,
                    (
                        code,
                        name,
                        phone,
                        email,
                        npwp,
                        pic,
                        address,
                        is_active,
                    ),
                    commit=True,
                )

                insert_count += 1

        messagebox.showinfo(
            "Import selesai",
            f"Pemasok baru : {insert_count}\nUpdate : {update_count}",
        )

        self.load_data()
