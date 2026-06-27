import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

from utils import parse_number, format_number, bind_number_entry

from openpyxl import Workbook
from tkinter import filedialog


def treeview_sort_column(tree, col, reverse=False, numeric=False):

    data = [(tree.set(k, col), k) for k in tree.get_children("")]

    if numeric:

        def to_num(v):
            try:
                v = str(v).replace(".", "").replace(",", ".")
                return float(v)
            except:
                return 0

        data.sort(key=lambda t: to_num(t[0]), reverse=reverse)

    else:
        data.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)

    for index, (_, k) in enumerate(data):
        tree.move(k, "", index)

    tree.heading(
        col, command=lambda: treeview_sort_column(tree, col, not reverse, numeric)
    )


class ProductView(tk.Frame):

    def __init__(self, parent, db, refresh_callback=None):

        super().__init__(parent, bg="#f9fafb")

        self.db = db
        self.refresh_callback = refresh_callback
        self.selected_id = None
        self.selected_products = set()

        self.build_ui()

    # ======================================
    # UI
    # ======================================

    def build_ui(self):

        header = tk.Frame(self, bg="#f9fafb")
        header.pack(fill="x")

        tk.Label(
            header, text="Master Produk", font=("Segoe UI", 20, "bold"), bg="#f9fafb"
        ).pack(side="left", padx=10, pady=10)

        toolbar = tk.Frame(self, bg="#f9fafb")
        toolbar.pack(fill="x", pady=5)

        # =========================
        # SEARCH & FILTER
        # =========================

        search_frame = tk.Frame(self, bg="#f9fafb")
        search_frame.pack(fill="x", pady=5)

        self.select_all_var = tk.BooleanVar()

        def toggle_select_all():
            checked = self.select_all_var.get()

            self.selected_products.clear()

            for item in self.tree.get_children():
                vals = self.tree.item(item, "values")
                sku = vals[2]  # karena kolom checkbox ditambah
                if checked:
                    self.selected_products.add(sku)

            self.refresh_checkboxes()

        tk.Checkbutton(
            search_frame,
            variable=self.select_all_var,
            command=toggle_select_all,
            bg="#f9fafb",
        ).pack(side="left", padx=(5, 5))

        tk.Label(search_frame, text="Cari Produk:", bg="#f9fafb").pack(
            side="left", padx=(5, 2)
        )

        self.search_entry = tk.Entry(search_frame, width=30)
        self.search_entry.pack(side="left", padx=5)

        self.search_entry.bind("<KeyRelease>", lambda e: self.load_data())

        # kategori
        tk.Label(search_frame, text="Kategori:", bg="#f9fafb").pack(
            side="left", padx=(15, 2)
        )

        cats = self.db.execute("SELECT name FROM categories ORDER BY name").fetchall()

        cat_list = ["SEMUA"] + [c["name"] for c in cats]

        self.category_filter = ttk.Combobox(
            search_frame, values=cat_list, width=20, state="readonly"
        )

        self.category_filter.set("SEMUA")
        self.category_filter.pack(side="left", padx=5)

        self.category_filter.bind("<<ComboboxSelected>>", lambda e: self.load_data())

        tk.Label(search_frame, text="Status:", bg="#f9fafb").pack(
            side="left", padx=(15, 2)
        )

        self.status_filter = ttk.Combobox(
            search_frame,
            values=["SEMUA", "AKTIF", "NONAKTIF"],
            width=12,
            state="readonly",
        )

        self.status_filter.set("AKTIF")
        self.status_filter.pack(side="left", padx=5)

        self.status_filter.bind("<<ComboboxSelected>>", lambda e: self.load_data())

        # TOTAL PRODUK
        self.total_label = tk.Label(
            search_frame,
            text="Total Produk: 0",
            bg="#f9fafb",
            font=("Segoe UI", 10, "bold"),
        )

        self.total_label.pack(side="right", padx=10)

        tk.Button(
            toolbar,
            text="Tambah Produk",
            bg="#10b981",
            width=14,
            command=self.popup_add_product,
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar,
            text="Edit Produk",
            bg="#f59e0b",
            width=14,
            command=self.update_product,
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar, text="Hapus", bg="#ef4444", width=14, command=self.delete_product
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar, text="Cetak", bg="#AE00FF", width=14, command=self.print_products
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar,
            text="Edit Kategori",
            bg="#D45800",
            width=14,
            command=self.popup_edit_category,
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar,
            text="Edit Satuan",
            bg="#00FF00",
            width=14,
            command=self.popup_edit_unit,
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar,
            text="Import XLSX",
            bg="#6366f1",
            width=14,
            command=self.import_products,
        ).pack(side="left", padx=5)

        tk.Button(
            toolbar, text="Refresh", bg="#3b82f6", width=14, command=self.load_data
        ).pack(side="right", padx=5)

        # =========================
        # TABLE
        # =========================

        self.tree = ttk.Treeview(
            self,
            columns=(
                "check",
                "no",
                "sku",
                "name",
                "supplier",
                "category",
                "unit",
                "cost",
                "sell",
                "stock",
                "min",
            ),
            show="headings",
        )

        headers = [
            ("check", "", 40, False),
            ("no", "No", 40, True),
            ("sku", "SKU", 120, False),
            ("name", "Nama", 220, False),
            ("supplier", "Supplier", 180, False),
            ("category", "Kategori", 140, False),
            ("unit", "Satuan", 100, False),
            ("cost", "Harga Beli", 120, True),
            ("sell", "Harga Jual", 120, True),
            ("stock", "Stok", 80, True),
            ("min", "Min", 80, True),
        ]

        for c, h, w, num in headers:

            self.tree.heading(
                c,
                text=h,
                command=lambda col=c, numeric=num: treeview_sort_column(
                    self.tree, col, False, numeric
                ),
            )

            self.tree.column(c, width=w)

        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Button-1>", self.toggle_checkbox)
        self.tree.tag_configure("inactive", foreground="red")

        self.load_data()

    def toggle_checkbox(self, event):

        region = self.tree.identify("region", event.x, event.y)

        if region != "cell":
            return

        column = self.tree.identify_column(event.x)

        if column != "#1":
            return

        item = self.tree.identify_row(event.y)

        if not item:
            return

        vals = list(self.tree.item(item, "values"))
        sku = vals[2]

        if sku in self.selected_products:
            self.selected_products.remove(sku)
        else:
            self.selected_products.add(sku)

        self.refresh_checkboxes()

    def refresh_checkboxes(self):

        for item in self.tree.get_children():

            vals = list(self.tree.item(item, "values"))
            sku = vals[2]

            if sku in self.selected_products:
                vals[0] = "☑"
            else:
                vals[0] = "☐"

            self.tree.item(item, values=vals)

    # ======================================
    # POPUP TAMBAH PRODUK
    # ======================================

    def popup_add_product(self):

        win = tk.Toplevel(self)
        win.title("Tambah Produk")
        win.geometry("600x350")

        frame = tk.Frame(win)
        frame.pack(padx=20, pady=20, fill="both", expand=True)

        tk.Label(frame, text="SKU").grid(row=0, column=0, sticky="w")
        sku_entry = tk.Entry(frame, width=30)
        sku_entry.grid(row=0, column=1, pady=5)

        tk.Label(frame, text="Nama").grid(row=1, column=0, sticky="w")
        name_entry = tk.Entry(frame, width=30)
        name_entry.grid(row=1, column=1, pady=5)

        tk.Label(frame, text="Supplier").grid(row=2, column=0, sticky="w")

        suppliers = self.db.execute(
            "SELECT id,name FROM suppliers ORDER BY name"
        ).fetchall()

        supplier_map = {r["name"]: r["id"] for r in suppliers}

        supplier_combo = ttk.Combobox(
            frame, values=list(supplier_map.keys()), state="readonly"
        )

        supplier_combo.grid(row=2, column=1, pady=5)

        tk.Label(frame, text="Kategori").grid(row=3, column=0, sticky="w")

        cats = self.db.execute("SELECT name FROM categories").fetchall()
        cat_list = [c["name"] for c in cats]

        cat_combo = ttk.Combobox(frame, values=cat_list, state="readonly")
        cat_combo.grid(row=3, column=1, pady=5)

        tk.Label(frame, text="Satuan").grid(row=4, column=0, sticky="w")

        units = self.db.execute("SELECT name FROM units").fetchall()
        unit_list = [u["name"] for u in units]

        unit_combo = ttk.Combobox(frame, values=unit_list, state="readonly")
        unit_combo.grid(row=4, column=1, pady=5)

        tk.Label(frame, text="Harga Beli").grid(row=0, column=2, padx=30)
        cost_entry = tk.Entry(frame, width=20)
        cost_entry.grid(row=0, column=3)
        bind_number_entry(cost_entry)

        tk.Label(frame, text="Harga Jual").grid(row=1, column=2, padx=30)
        sell_entry = tk.Entry(frame, width=20)
        sell_entry.grid(row=1, column=3)
        bind_number_entry(sell_entry)

        tk.Label(frame, text="Min Stok").grid(row=2, column=2, padx=30)
        min_entry = tk.Entry(frame, width=20)
        min_entry.grid(row=2, column=3)
        bind_number_entry(min_entry)

        # ======================
        # SAVE
        # ======================

        def save():

            sku = sku_entry.get().strip()
            name = name_entry.get().strip()

            supplier = supplier_combo.get()
            cat = cat_combo.get()
            unit = unit_combo.get()

            cost = parse_number(cost_entry.get())
            sell = parse_number(sell_entry.get())
            min_stock = parse_number(min_entry.get())

            if not sku or not name:
                messagebox.showerror("Error", "SKU dan Nama wajib")
                return

            supplier_id = supplier_map.get(supplier)

            cat_id = self.get_id("categories", cat)
            unit_id = self.get_id("units", unit)

            self.db.execute(
                """
            INSERT INTO products
            (sku,name,supplier_id,category_id,unit_id,
            cost_price,sell_price,min_stock)
            VALUES(?,?,?,?,?,?,?,?)
            """,
                (sku, name, supplier_id, cat_id, unit_id, cost, sell, min_stock),
                commit=True,
            )

            messagebox.showinfo("Sukses", "Produk disimpan")

            win.destroy()

            self.load_data()

        tk.Button(frame, text="Simpan", bg="#06b6d4", width=14, command=save).grid(
            row=6, column=3, pady=20
        )

    # ======================================
    # LOAD DATA
    # ======================================

    def load_data(self):

        # hitung total produk
        row = self.db.execute("SELECT COUNT(*) total FROM products").fetchone()

        total_produk = row["total"] if row else 0

        self.total_label.config(text=f"Total Produk : {total_produk} item")

        for i in self.tree.get_children():
            self.tree.delete(i)

        keyword = ""
        kategori = ""
        status_filter = "SEMUA"

        if hasattr(self, "search_entry"):
            keyword = self.search_entry.get().lower()

        if hasattr(self, "category_filter"):
            kategori = self.category_filter.get()

        if hasattr(self, "status_filter"):
            status_filter = self.status_filter.get()

        rows = self.db.execute(
            """
        SELECT
        p.id,
        p.sku,
        p.name,
        s.name supplier,
        c.name cat,
        u.name unit,
        p.cost_price,
        p.sell_price,
        p.stock,
        p.min_stock,
        p.is_active
        FROM products p
        LEFT JOIN suppliers s ON s.id=p.supplier_id
        LEFT JOIN categories c ON c.id=p.category_id
        LEFT JOIN units u ON u.id=p.unit_id
        ORDER BY p.name ASC
        """
        ).fetchall()

        no = 1

        for r in rows:

            name = (r["name"] or "").lower()
            sku = (r["sku"] or "").lower()
            cat = r["cat"] or ""

            # filter search
            if keyword:
                if keyword not in name and keyword not in sku:
                    continue

            # filter kategori
            if kategori and kategori != "SEMUA":
                if cat != kategori:
                    continue

            if status_filter == "AKTIF" and r["is_active"] == 0:
                continue

            if status_filter == "NONAKTIF" and r["is_active"] == 1:
                continue

            tag = ""
            name = r["name"]

            if r["is_active"] == 0:
                name = "✖ " + name
                tag = "inactive"

            self.tree.insert(
                "",
                "end",
                values=(
                    "☐",
                    no,
                    r["sku"],
                    name,
                    r["supplier"],
                    r["cat"],
                    r["unit"],
                    format_number(r["cost_price"]),
                    format_number(r["sell_price"]),
                    format_number(r["stock"]),
                    format_number(r["min_stock"]),
                ),
                tags=(tag,),
            )

            no += 1

    # ======================================
    # SELECT
    # ======================================

    def on_select(self, event=None):

        s = self.tree.selection()

        if not s:
            return

        vals = self.tree.item(s[0], "values")

        # SKU ada di kolom index 2
        sku = vals[2]

        row = self.db.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone()

        if row:
            self.selected_id = row["id"]

    # ======================================
    # UPDATE
    # ======================================

    def update_product(self):

        if not self.selected_id:
            messagebox.showerror("Error", "Pilih produk terlebih dahulu")
            return

        row = self.db.execute(
            """
            SELECT *
            FROM products
            WHERE id=?
        """,
            (self.selected_id,),
        ).fetchone()

        if not row:
            return

        win = tk.Toplevel(self)
        win.title("Edit Produk")
        win.geometry("600x350")

        frame = tk.Frame(win)
        frame.pack(padx=20, pady=20, fill="both", expand=True)

        tk.Label(frame, text="SKU").grid(row=0, column=0, sticky="w")
        sku_entry = tk.Entry(frame, width=30)
        sku_entry.grid(row=0, column=1, pady=5)

        tk.Label(frame, text="Nama").grid(row=1, column=0, sticky="w")
        name_entry = tk.Entry(frame, width=30)
        name_entry.grid(row=1, column=1, pady=5)

        tk.Label(frame, text="Supplier").grid(row=2, column=0, sticky="w")

        suppliers = self.db.execute(
            "SELECT id,name FROM suppliers ORDER BY name"
        ).fetchall()

        supplier_map = {r["name"]: r["id"] for r in suppliers}

        supplier_combo = ttk.Combobox(
            frame, values=list(supplier_map.keys()), state="readonly"
        )

        supplier_combo.grid(row=2, column=1, pady=5)

        tk.Label(frame, text="Kategori").grid(row=3, column=0, sticky="w")

        cats = self.db.execute("SELECT name FROM categories").fetchall()
        cat_list = [c["name"] for c in cats]

        cat_combo = ttk.Combobox(frame, values=cat_list, state="readonly")
        cat_combo.grid(row=3, column=1, pady=5)

        tk.Label(frame, text="Satuan").grid(row=4, column=0, sticky="w")

        units = self.db.execute("SELECT name FROM units").fetchall()
        unit_list = [u["name"] for u in units]

        unit_combo = ttk.Combobox(frame, values=unit_list, state="readonly")
        unit_combo.grid(row=4, column=1, pady=5)

        tk.Label(frame, text="Harga Beli").grid(row=0, column=2, padx=30)
        cost_entry = tk.Entry(frame, width=20)
        cost_entry.grid(row=0, column=3)
        bind_number_entry(cost_entry)

        tk.Label(frame, text="Harga Jual").grid(row=1, column=2, padx=30)
        sell_entry = tk.Entry(frame, width=20)
        sell_entry.grid(row=1, column=3)
        bind_number_entry(sell_entry)

        tk.Label(frame, text="Min Stok").grid(row=2, column=2, padx=30)
        min_entry = tk.Entry(frame, width=20)
        min_entry.grid(row=2, column=3)
        bind_number_entry(min_entry)

        active_var = tk.IntVar(value=row["is_active"])

        active_check = tk.Checkbutton(
            frame, text="Nonaktifkan Produk", variable=active_var, onvalue=0, offvalue=1
        )

        active_check.grid(row=3, column=3)

        # =========================
        # ISI DATA PRODUK
        # =========================

        sku_entry.insert(0, row["sku"])
        name_entry.insert(0, row["name"])
        cost_entry.insert(0, row["cost_price"])
        sell_entry.insert(0, row["sell_price"])
        min_entry.insert(0, row["min_stock"])

        # supplier
        supp = self.db.execute(
            "SELECT name FROM suppliers WHERE id=?", (row["supplier_id"],)
        ).fetchone()

        if supp:
            supplier_combo.set(supp["name"])

        # kategori
        cat = self.db.execute(
            "SELECT name FROM categories WHERE id=?", (row["category_id"],)
        ).fetchone()

        if cat:
            cat_combo.set(cat["name"])

        # satuan
        unit = self.db.execute(
            "SELECT name FROM units WHERE id=?", (row["unit_id"],)
        ).fetchone()

        if unit:
            unit_combo.set(unit["name"])

        # =========================
        # UPDATE DATA
        # =========================

        def save():

            sku = sku_entry.get().strip()
            name = name_entry.get().strip()

            supplier = supplier_combo.get()
            cat = cat_combo.get()
            unit = unit_combo.get()

            cost = parse_number(cost_entry.get())
            sell = parse_number(sell_entry.get())
            min_stock = parse_number(min_entry.get())

            supplier_id = supplier_map.get(supplier)
            cat_id = self.get_id("categories", cat)
            unit_id = self.get_id("units", unit)

            self.db.execute(
                """
            UPDATE products
            SET sku=?,name=?,supplier_id=?,category_id=?,unit_id=?,
                cost_price=?,sell_price=?,min_stock=?,is_active=?
            WHERE id=?
            """,
                (
                    sku,
                    name,
                    supplier_id,
                    cat_id,
                    unit_id,
                    cost,
                    sell,
                    min_stock,
                    active_var.get(),
                    self.selected_id,
                ),
                commit=True,
            )

            messagebox.showinfo("Sukses", "Produk berhasil diupdate")

            win.destroy()
            self.load_data()

        tk.Button(
            frame, text="Update Produk", bg="#f59e0b", width=14, command=save
        ).grid(row=6, column=3, pady=20)

    # ======================================
    # DELETE
    # ======================================

    def delete_product(self):

        # jika ada checkbox dipilih
        if self.selected_products:
            targets = list(self.selected_products)

        else:
            # jika tidak ada checkbox gunakan baris yang dipilih
            sel = self.tree.selection()

            if not sel:
                messagebox.showerror("Error", "Pilih atau centang produk dulu")
                return

            vals = self.tree.item(sel[0], "values")
            targets = [vals[2]]  # SKU ada di kolom ke-3

        win = tk.Toplevel(self)
        win.title("Konfirmasi")
        win.geometry("350x150")
        win.configure(bg="#e5e7eb")

        tk.Label(
            win,
            text="Hapus Produk?",
            font=("Segoe UI", 14, "bold"),
            bg="#e5e7eb",
        ).pack(pady=20)

        btn_frame = tk.Frame(win, bg="#e5e7eb")
        btn_frame.pack()

        def confirm_delete():

            for sku in targets:

                row = self.db.execute(
                    "SELECT id FROM products WHERE sku=?", (sku,)
                ).fetchone()

                if row:
                    self.db.execute(
                        "DELETE FROM products WHERE id=?", (row["id"],), commit=True
                    )

            self.selected_products.clear()

            win.destroy()

            messagebox.showinfo("Sukses", "Produk berhasil dihapus")

            self.load_data()

        tk.Button(
            btn_frame,
            text="Ya",
            width=10,
            bg="#16a34a",
            fg="white",
            command=confirm_delete,
        ).pack(side="left", padx=10)

        tk.Button(
            btn_frame,
            text="Tidak",
            width=10,
            bg="#dc2626",
            fg="white",
            command=win.destroy,
        ).pack(side="left", padx=10)

    # ======================================
    # CETAK PRODUK
    # ======================================

    def print_products(self):

        win = tk.Toplevel(self)
        win.title("Konfirmasi")
        win.geometry("350x150")
        win.configure(bg="#e5e7eb")

        tk.Label(
            win, text="Cetak Produk?", font=("Segoe UI", 14, "bold"), bg="#e5e7eb"
        ).pack(pady=20)

        btn_frame = tk.Frame(win, bg="#e5e7eb")
        btn_frame.pack()

        def confirm_print():

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")],
                title="Simpan File Produk",
            )

            if not filepath:
                return

            rows = self.db.execute(
                """
            SELECT
            p.sku,
            p.name,
            s.name supplier,
            c.name category,
            u.name unit,
            p.cost_price,
            p.sell_price,
            p.stock,
            p.min_stock,
            p.is_active
            FROM products p
            LEFT JOIN suppliers s ON s.id=p.supplier_id
            LEFT JOIN categories c ON c.id=p.category_id
            LEFT JOIN units u ON u.id=p.unit_id
            ORDER BY p.name
            """
            ).fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = "Produk"

            headers = [
                "SKU",
                "Nama Produk",
                "Supplier",
                "Kategori",
                "Satuan",
                "Harga Beli",
                "Harga Jual",
                "Stok",
                "Min Stok",
                "status",
            ]

            ws.append(headers)

            for r in rows:

                status = "AKTIF" if r["is_active"] == 1 else "NONAKTIF"

                ws.append(
                    [
                        r["sku"],
                        r["name"],
                        r["supplier"],
                        r["category"],
                        r["unit"],
                        r["cost_price"],
                        r["sell_price"],
                        r["stock"],
                        r["min_stock"],
                        status,
                    ]
                )

            wb.save(filepath)

            win.destroy()

            messagebox.showinfo("Sukses", "Produk berhasil dicetak ke Excel")

        tk.Button(
            btn_frame,
            text="Ya",
            width=10,
            bg="#16a34a",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=confirm_print,
        ).pack(side="left", padx=10)

        tk.Button(
            btn_frame,
            text="Tidak",
            width=10,
            bg="#dc2626",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=win.destroy,
        ).pack(side="left", padx=10)

    def import_products(self):

        from openpyxl import load_workbook
        from tkinter import filedialog

        filepath = filedialog.askopenfilename(
            filetypes=[("Excel File", "*.xlsx")], title="Pilih File Produk"
        )

        if not filepath:
            return

        try:

            wb = load_workbook(filepath)
            ws = wb.active

            insert_count = 0
            update_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):

                sku = row[0]
                name = row[1]
                supplier = row[2]
                category = row[3]
                unit = row[4]
                cost = row[5] or 0
                sell = row[6] or 0
                stock = row[7] or 0
                min_stock = row[8] or 0
                status = row[9] if len(row) > 9 else "AKTIF"

                is_active = 1
                if str(status).upper() == "NONAKTIF":
                    is_active = 0

                if not sku or not name:
                    continue

                supplier_id = self.get_id("suppliers", supplier)
                category_id = self.get_id("categories", category)
                unit_id = self.get_id("units", unit)

                # cek produk berdasarkan SKU
                existing = self.db.execute(
                    "SELECT id FROM products WHERE sku=?", (sku,)
                ).fetchone()

                if existing:

                    # UPDATE
                    self.db.execute(
                        """
                        UPDATE products
                        SET
                            name=?,
                            supplier_id=?,
                            category_id=?,
                            unit_id=?,
                            cost_price=?,
                            sell_price=?,
                            stock=?,
                            min_stock=?,
                            is_active=?
                        WHERE sku=?
                        """,
                        (
                            name,
                            supplier_id,
                            category_id,
                            unit_id,
                            cost,
                            sell,
                            stock,
                            min_stock,
                            is_active,
                            sku,
                        ),
                        commit=True,
                    )

                    update_count += 1

                else:

                    # INSERT
                    self.db.execute(
                        """
                        INSERT INTO products
                        (sku,name,supplier_id,category_id,unit_id,
                        cost_price,sell_price,stock,min_stock,is_active)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            sku,
                            name,
                            supplier_id,
                            category_id,
                            unit_id,
                            cost,
                            sell,
                            stock,
                            min_stock,
                            is_active,
                        ),
                        commit=True,
                    )

                    insert_count += 1

            messagebox.showinfo(
                "Import Selesai",
                f"""
    Produk baru : {insert_count}
    Produk update : {update_count}
    """,
            )

            self.load_data()

        except Exception as e:

            messagebox.showerror("Error Import", str(e))

    # ======================================
    # GET ID
    # ======================================

    def get_id(self, table, name):

        if not name:
            return None

        r = self.db.execute(f"SELECT id FROM {table} WHERE name=?", (name,)).fetchone()

        return r["id"] if r else None

    def popup_edit_category(self):

        CategoryEditor(self, self.db, self.load_data)

    def popup_edit_unit(self):

        UnitEditor(self, self.db, self.load_data)


class CategoryEditor(tk.Toplevel):

    def __init__(self, parent, db, refresh_callback=None):
        super().__init__(parent)

        self.db = db
        self.refresh_callback = refresh_callback
        self.selected_id = None

        self.title("Edit Kategori")
        self.geometry("600x450")

        self.build_ui()
        self.load_data()

    def build_ui(self):

        frame = tk.Frame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        # =============================
        # INPUT
        # =============================

        tk.Label(frame, text="Nama Kategori").grid(row=0, column=0, sticky="w")

        self.name_entry = tk.Entry(frame, width=40)
        self.name_entry.grid(row=0, column=1, padx=10)

        tk.Button(
            frame, text="Save / Add", bg="#22c1dc", command=self.save_category
        ).grid(row=0, column=2, padx=5)

        tk.Button(
            frame, text="Hapus", bg="#dc2626", fg="white", command=self.delete_category
        ).grid(row=0, column=3, padx=5)

        # =============================
        # TABLE
        # =============================

        self.tree = ttk.Treeview(
            frame, columns=("no", "name"), show="headings", height=12
        )

        headers = [("no", "No", 60, True), ("name", "Kategori", 350, False)]

        for c, h, w, num in headers:

            self.tree.heading(
                c,
                text=h,
                command=lambda col=c, numeric=num: treeview_sort_column(
                    self.tree, col, False, numeric
                ),
            )

            self.tree.column(c, width=w)

        self.tree.column("no", width=60)
        self.tree.column("name", width=350)

        self.tree.grid(row=1, column=0, columnspan=4, pady=10, sticky="nsew")

        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(1, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        # =============================
        # BUTTONS BAWAH
        # =============================

        bottom = tk.Frame(frame)
        bottom.grid(row=2, column=0, columnspan=4, pady=10)

        tk.Button(
            bottom,
            text="Tutup",
            bg="#dc2626",
            fg="white",
            width=12,
            command=self.destroy,
        ).pack(side="left", padx=5)

    # =============================
    # LOAD DATA
    # =============================

    def load_data(self):

        for i in self.tree.get_children():
            self.tree.delete(i)

        rows = self.db.execute(
            "SELECT id,name FROM categories ORDER BY name"
        ).fetchall()

        no = 1

        for r in rows:

            self.tree.insert("", "end", values=(no, r["name"]), tags=(r["id"],))

            no += 1

    # =============================
    # SELECT
    # =============================

    def on_select(self, event=None):

        s = self.tree.selection()

        if not s:
            return

        item = self.tree.item(s[0])

        self.selected_id = item["tags"][0]

        vals = item["values"]

        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, vals[1])

    # =============================
    # SAVE / UPDATE
    # =============================

    def save_category(self):

        name = self.name_entry.get().strip()

        if not name:
            messagebox.showerror("Error", "Nama kategori kosong")
            return

        if self.selected_id:

            self.db.execute(
                "UPDATE categories SET name=? WHERE id=?",
                (name, self.selected_id),
                commit=True,
            )

        else:

            self.db.execute(
                "INSERT INTO categories(name) VALUES(?)", (name,), commit=True
            )

        self.name_entry.delete(0, "end")
        self.selected_id = None

        self.load_data()

        if self.refresh_callback:
            self.refresh_callback()

    # =============================
    # DELETE
    # =============================

    def delete_category(self):

        if not self.selected_id:
            return

        if not messagebox.askyesno("Konfirmasi", "Hapus kategori?"):
            return

        self.db.execute(
            "DELETE FROM categories WHERE id=?", (self.selected_id,), commit=True
        )

        self.selected_id = None
        self.name_entry.delete(0, "end")

        self.load_data()

        if self.refresh_callback:
            self.refresh_callback()


class UnitEditor(tk.Toplevel):

    def __init__(self, parent, db, refresh_callback=None):
        super().__init__(parent)

        self.db = db
        self.refresh_callback = refresh_callback
        self.selected_id = None

        self.title("Edit Satuan")
        self.geometry("600x450")

        self.build_ui()
        self.load_data()

    def build_ui(self):

        frame = tk.Frame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        tk.Label(frame, text="Nama Satuan").grid(row=0, column=0, sticky="w")

        self.name_entry = tk.Entry(frame, width=40)
        self.name_entry.grid(row=0, column=1, padx=10)

        tk.Button(frame, text="Save / Add", bg="#22c1dc", command=self.save_unit).grid(
            row=0, column=2, padx=5
        )

        tk.Button(
            frame, text="Hapus", bg="#dc2626", fg="white", command=self.delete_unit
        ).grid(row=0, column=3, padx=5)

        # ======================
        # TABLE
        # ======================

        self.tree = ttk.Treeview(
            frame, columns=("no", "name"), show="headings", height=12
        )

        headers = [("no", "No", 60, True), ("name", "Satuan", 350, False)]

        for c, h, w, num in headers:

            self.tree.heading(
                c,
                text=h,
                command=lambda col=c, numeric=num: treeview_sort_column(
                    self.tree, col, False, numeric
                ),
            )

            self.tree.column(c, width=w)

        self.tree.grid(row=1, column=0, columnspan=4, pady=10, sticky="nsew")

        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(1, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        bottom = tk.Frame(frame)
        bottom.grid(row=2, column=0, columnspan=4, pady=10)

        tk.Button(
            bottom,
            text="Tutup",
            bg="#dc2626",
            fg="white",
            width=12,
            command=self.destroy,
        ).pack(side="left", padx=5)

    # ======================
    # LOAD DATA
    # ======================

    def load_data(self):

        for i in self.tree.get_children():
            self.tree.delete(i)

        rows = self.db.execute("SELECT id,name FROM units ORDER BY name").fetchall()

        no = 1

        for r in rows:

            self.tree.insert("", "end", values=(no, r["name"]), tags=(r["id"],))

            no += 1

    # ======================
    # SELECT
    # ======================

    def on_select(self, event=None):

        s = self.tree.selection()

        if not s:
            return

        item = self.tree.item(s[0])

        self.selected_id = item["tags"][0]

        vals = item["values"]

        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, vals[1])

    # ======================
    # SAVE / UPDATE
    # ======================

    def save_unit(self):

        name = self.name_entry.get().strip()

        if not name:
            messagebox.showerror("Error", "Nama satuan kosong")
            return

        if self.selected_id:

            self.db.execute(
                "UPDATE units SET name=? WHERE id=?",
                (name, self.selected_id),
                commit=True,
            )

        else:

            self.db.execute("INSERT INTO units(name) VALUES(?)", (name,), commit=True)

        self.name_entry.delete(0, "end")
        self.selected_id = None

        self.load_data()

        if self.refresh_callback:
            self.refresh_callback()

    # ======================
    # DELETE
    # ======================

    def delete_unit(self):

        if not self.selected_id:
            return

        if not messagebox.askyesno("Konfirmasi", "Hapus satuan?"):
            return

        self.db.execute(
            "DELETE FROM units WHERE id=?", (self.selected_id,), commit=True
        )

        self.selected_id = None
        self.name_entry.delete(0, "end")

        self.load_data()

        if self.refresh_callback:
            self.refresh_callback()


# =====================================================
# GENERIC MASTER VIEW
# =====================================================


class GenericMasterView(tk.Frame):

    def __init__(self, parent, db, title, table, code_prefix):

        super().__init__(parent, bg="#f9fafb")

        self.db = db
        self.title = title
        self.table = table
        self.code_prefix = code_prefix
        self.selected_id = None

        self.build_ui()

    def build_ui(self):

        tk.Label(
            self, text=self.title, font=("Segoe UI", 16, "bold"), bg="#f9fafb"
        ).pack(anchor="w", pady=(0, 10))

        form = tk.Frame(self, bg="white", bd=1, relief="solid")
        form.pack(fill="x", pady=5)

        tk.Label(form, text="Nama", bg="white").grid(
            row=0, column=0, padx=10, pady=(10, 2)
        )
        tk.Label(form, text="Telepon", bg="white").grid(
            row=0, column=1, padx=10, pady=(10, 2)
        )
        tk.Label(form, text="Alamat", bg="white").grid(
            row=0, column=2, padx=10, pady=(10, 2)
        )

        self.name_entry = tk.Entry(form)
        self.phone_entry = tk.Entry(form)
        self.addr_entry = tk.Entry(form)

        self.name_entry.grid(row=1, column=0, padx=10, pady=(0, 10))
        self.phone_entry.grid(row=1, column=1, padx=10, pady=(0, 10))
        self.addr_entry.grid(row=1, column=2, padx=10, pady=(0, 10))

        btn = tk.Frame(form, bg="white")
        btn.grid(row=2, column=0, columnspan=3, padx=10, pady=(0, 10), sticky="w")

        tk.Button(
            btn, text="Simpan", command=self.save_data, bg="#2563eb", fg="white"
        ).pack(side="left", padx=5)

        tk.Button(
            btn, text="Update", command=self.update_data, bg="#f59e0b", fg="white"
        ).pack(side="left", padx=5)

        tk.Button(
            btn, text="Hapus", command=self.delete_data, bg="#dc2626", fg="white"
        ).pack(side="left", padx=5)

        self.tree = ttk.Treeview(
            self,
            columns=("id", "code", "name", "phone", "address"),
            show="headings",
            height=18,
        )

        headers = [
            ("id", "ID", 50),
            ("code", "Kode", 100),
            ("name", "Nama", 220),
            ("phone", "Telepon", 140),
            ("address", "Alamat", 300),
        ]

        for c, h, w in headers:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w)

        self.tree.pack(fill="both", expand=True, pady=10)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        self.load_data()

    def load_data(self):

        for i in self.tree.get_children():
            self.tree.delete(i)

        rows = self.db.execute(
            f"SELECT * FROM {self.table} ORDER BY id DESC"
        ).fetchall()

        for r in rows:

            self.tree.insert(
                "",
                "end",
                values=(r["id"], r["code"], r["name"], r["phone"], r["address"]),
            )

    def on_select(self, event):

        s = self.tree.selection()

        if not s:
            return

        vals = self.tree.item(s[0], "values")

        self.selected_id = vals[0]

        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, vals[2])

        self.phone_entry.delete(0, "end")
        self.phone_entry.insert(0, vals[3])

        self.addr_entry.delete(0, "end")
        self.addr_entry.insert(0, vals[4])

    def save_data(self):

        name = self.name_entry.get()
        phone = self.phone_entry.get()
        address = self.addr_entry.get()

        count = (
            self.db.execute(f"SELECT COUNT(*) cnt FROM {self.table}").fetchone()["cnt"]
            + 1
        )

        code = f"{self.code_prefix}-{str(count).zfill(4)}"

        self.db.execute(
            f"INSERT INTO {self.table}(code,name,phone,address) VALUES(?,?,?,?)",
            (code, name, phone, address),
            commit=True,
        )

        self.load_data()

    def update_data(self):

        if not self.selected_id:
            return

        self.db.execute(
            f"UPDATE {self.table} SET name=?,phone=?,address=? WHERE id=?",
            (
                self.name_entry.get(),
                self.phone_entry.get(),
                self.addr_entry.get(),
                self.selected_id,
            ),
            commit=True,
        )

        self.load_data()

    def delete_data(self):

        if not self.selected_id:
            return

        self.db.execute(
            f"DELETE FROM {self.table} WHERE id=?", (self.selected_id,), commit=True
        )

        self.load_data()
