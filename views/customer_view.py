import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from tkinter import filedialog

from utils import treeview_sort_column
from utils import parse_number, bind_number_entry, format_number


class CustomerView(tk.Frame):

    def __init__(self, parent, db):
        super().__init__(parent, bg="#f9fafb")

        self.db = db
        self.selected_id = None

        self.build_ui()
        self.selected_customers = set()

    # ======================================
    # UI
    # ======================================

    def build_ui(self):

        header = tk.Frame(self, bg="#f9fafb")
        header.pack(fill="x")

        tk.Label(
            header, text="Master Pelanggan", font=("Segoe UI", 20, "bold"), bg="#f9fafb"
        ).pack(side="left", padx=10, pady=10)

        # ======================
        # BUTTON
        # ======================

        btn_frame = tk.Frame(self, bg="#f9fafb")
        btn_frame.pack(fill="x", pady=5)

        tk.Button(
            btn_frame,
            text="Tambah Pelanggan",
            bg="#10b981",
            fg="black",
            width=16,
            command=self.popup_add,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Edit Pelanggan",
            bg="#f59e0b",
            fg="black",
            width=16,
            command=self.popup_edit,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Hapus Pelanggan",
            bg="#ef4444",
            fg="black",
            width=16,
            command=self.delete_customer,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Cetak",
            bg="#8b5cf6",
            fg="black",
            width=16,
            command=self.print_customers,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Import XLSX",
            bg="#6366f1",
            fg="black",
            width=16,
            command=self.import_customers,
        ).pack(side="left", padx=5)

        tk.Button(
            btn_frame,
            text="Refresh",
            bg="#3b82f6",
            fg="black",
            width=16,
            command=self.load_data,
        ).pack(side="right", padx=5)

        # ======================
        # SEARCH
        # ======================

        search_frame = tk.Frame(self, bg="#f9fafb")
        search_frame.pack(fill="x", pady=5)

        # ======================
        # CHECKBOX ALL
        # ======================

        self.select_all_var = tk.BooleanVar()

        def toggle_select_all():

            checked = self.select_all_var.get()

            self.selected_customers.clear()

            if checked:

                for item in self.tree.get_children():

                    vals = self.tree.item(item, "values")

                    if not vals:
                        continue

                    code = vals[2]

                    self.selected_customers.add(code)

            self.refresh_checkboxes()

        tk.Checkbutton(
            search_frame,
            text="All",
            variable=self.select_all_var,
            command=toggle_select_all,
            bg="#f9fafb",
        ).pack(side="left", padx=(10, 5))

        # ======================
        # SEARCH
        # ======================

        tk.Label(search_frame, text="Cari pelanggan:", bg="#f9fafb").pack(
            side="left", padx=(5, 2)
        )

        self.search_entry = tk.Entry(search_frame, width=30)
        self.search_entry.pack(side="left", padx=5)

        self.search_entry.bind("<KeyRelease>", lambda e: self.load_data())

        # ======================
        # STATUS (dipindah ke kanan)
        # ======================

        tk.Label(search_frame, text="Status:", bg="#f9fafb").pack(
            side="left", padx=(15, 2)
        )

        self.status_filter = ttk.Combobox(
            search_frame,
            values=["SEMUA", "AKTIF", "NONAKTIF"],
            width=12,
            state="readonly",
        )

        self.status_filter.set("AKTIF")
        self.status_filter.pack(side="left", padx=5)

        self.status_filter.bind("<<ComboboxSelected>>", lambda e: self.load_data())

        # ======================
        # TABLE
        # ======================

        self.tree = ttk.Treeview(
            self,
            columns=(
                "check",
                "no",
                "code",
                "name",
                "phone",
                "email",
                "npwp",
                "sales",
                "domisili",
                "address",
                "credit",
                "saldo",
                "status",
            ),
            show="headings",
            height=20,
        )

        headers = [
            ("check", "", 40),
            ("no", "No", 60),
            ("code", "Kode", 120),
            ("name", "Nama", 200),
            ("phone", "Telepon", 140),
            ("email", "Email", 200),
            ("npwp", "NPWP", 160),
            ("sales", "Sales PIC", 120),
            ("domisili", "Domisili", 120),
            ("address", "Alamat", 120),
            ("credit", "Limit Kredit", 140),
            ("saldo", "Saldo", 140),
            ("status", "Status", 100),
        ]

        for c, h, w in headers:

            # kolom angka
            numeric_cols = ["credit", "saldo"]

            if c in numeric_cols:

                self.tree.heading(
                    c,
                    text=h,
                    command=lambda col=c: treeview_sort_column(
                        self.tree, col, False, True
                    ),
                )

            else:

                self.tree.heading(
                    c,
                    text=h,
                    command=lambda col=c: treeview_sort_column(self.tree, col, False),
                )

            self.tree.column(c, width=w)

        self.tree.tag_configure("inactive", background="#f3f4f6", foreground="#6b7280")

        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Button-1>", self.toggle_checkbox)

        self.load_data()

    # ======================================
    # LOAD DATA
    # ======================================

    def load_data(self):

        for i in self.tree.get_children():
            self.tree.delete(i)

        keyword = self.search_entry.get().lower()

        rows = self.db.execute(
            """
            SELECT
                c.*,
                COALESCE(SUM(si.total - si.paid),0) saldo,
                COALESCE(SUM(
                    CASE WHEN si.status!='PAID'
                    THEN si.total - si.paid
                    ELSE 0 END
                ),0) piutang
            FROM customers c
            LEFT JOIN sales_invoices si
                ON si.customer_id = c.id
            GROUP BY c.id
            ORDER BY c.name
            """
        ).fetchall()
        status_filter = (
            self.status_filter.get() if hasattr(self, "status_filter") else "SEMUA"
        )

        no = 1

        for r in rows:

            if status_filter == "AKTIF" and r["is_active"] == 0:
                continue

            if status_filter == "NONAKTIF" and r["is_active"] == 1:
                continue

            # HITUNG SALDO PELANGGAN

            name = (r["name"] or "").lower()

            if keyword and keyword not in name:
                continue

            name = r["name"]

            if r["is_active"] == 0:
                name = "✖ " + name
                tag = "inactive"
            else:
                tag = ""

            saldo = r["saldo"] or 0
            receivable = r["piutang"] or 0
            credit_limit = r["credit_limit"] or 0

            tags = []

            if credit_limit > 0 and receivable > credit_limit:
                tags.append("limit")

            if r["is_active"] == 0:
                tags.append("inactive")

            status = "Aktif" if r["is_active"] == 1 else "Non Aktif"

            self.tree.insert(
                "",
                "end",
                values=(
                    "☐",
                    no,
                    r["code"],
                    name,
                    r["phone"],
                    r["email"],
                    r["npwp"],
                    r["sales_pic"],
                    r["domisili"],
                    r["address"],
                    format_number(r["credit_limit"]),
                    format_number(saldo),
                    status,
                ),
                tags=(tag,),
            )

            no += 1

    # ======================================
    # SELECT
    # ======================================

    def on_select(self, event=None):

        s = self.tree.selection()

        if not s:
            self.selected_id = None
            return

        vals = self.tree.item(s[0], "values")

        code = vals[2]  # ✅ kolom kode pelanggan

        row = self.db.execute(
            "SELECT id FROM customers WHERE code=?", (code,)
        ).fetchone()

        if row:
            self.selected_id = row["id"]

    # ======================================
    # ADD
    # ======================================

    def popup_add(self):

        win = tk.Toplevel(self)
        win.title("Tambah Pelanggan")
        win.geometry("620x320")

        frame = tk.Frame(win)
        frame.pack(padx=20, pady=20, fill="both", expand=True)

        # ========================
        # LEFT SIDE
        # ========================

        tk.Label(frame, text="Nama").grid(row=0, column=0, sticky="w")
        name_entry = tk.Entry(frame, width=30)
        name_entry.grid(row=0, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Telepon").grid(row=1, column=0, sticky="w")
        phone_entry = tk.Entry(frame, width=30)
        phone_entry.grid(row=1, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Email").grid(row=2, column=0, sticky="w")
        email_entry = tk.Entry(frame, width=30)
        email_entry.grid(row=2, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="NPWP").grid(row=3, column=0, sticky="w")
        npwp_entry = tk.Entry(frame, width=30)
        npwp_entry.grid(row=3, column=1, pady=5, padx=(0, 20))

        # ========================
        # RIGHT SIDE
        # ========================

        tk.Label(frame, text="Limit Kredit").grid(row=0, column=2, sticky="w")
        credit_entry = tk.Entry(frame, width=30)
        credit_entry.grid(row=0, column=3, pady=5)

        bind_number_entry(credit_entry)

        tk.Label(frame, text="Sales PIC").grid(row=1, column=2, sticky="w")
        sales_entry = tk.Entry(frame, width=30)
        sales_entry.grid(row=1, column=3, pady=5)

        tk.Label(frame, text="Domisili").grid(row=2, column=2, sticky="w")
        dom_entry = tk.Entry(frame, width=30)
        dom_entry.grid(row=2, column=3, pady=5)

        tk.Label(frame, text="Alamat").grid(row=3, column=2, sticky="w")
        addr_entry = tk.Entry(frame, width=30)
        addr_entry.grid(row=3, column=3, pady=5)

        # ========================
        # SAVE
        # ========================

        def save():

            name = name_entry.get().strip()
            phone = phone_entry.get().strip()
            email = email_entry.get().strip()
            npwp = npwp_entry.get().strip()

            credit = parse_number(credit_entry.get())
            sales = sales_entry.get().strip()
            dom = dom_entry.get().strip()
            addr = addr_entry.get().strip()

            if not name:
                messagebox.showerror("Error", "Nama pelanggan wajib diisi")
                return

            # ==============================
            # GENERATE CUSTOMER CODE
            # ==============================

            row = self.db.execute("SELECT MAX(id) max_id FROM customers").fetchone()

            next_id = (row["max_id"] or 0) + 1

            code = f"CUST-{str(next_id).zfill(4)}"

            # ==============================
            # INSERT DATABASE
            # ==============================

            self.db.execute(
                """
                INSERT INTO customers
                (code,name,phone,email,npwp,credit_limit,sales_pic,domisili,address)
                VALUES(?,?,?,?,?,?,?,?,?)
                """,
                (
                    code,
                    name,
                    phone,
                    email,
                    npwp,
                    credit,
                    sales,
                    dom,
                    addr,
                ),
                commit=True,
            )

            messagebox.showinfo("Sukses", "Pelanggan berhasil disimpan")

            win.destroy()

            self.load_data()

        tk.Button(frame, text="Simpan", bg="#06b6d4", width=14, command=save).grid(
            row=5, column=3, pady=20
        )

    # ======================================
    # EDIT
    # ======================================

    def popup_edit(self):

        if not self.selected_id:
            messagebox.showerror("Error", "Pilih pelanggan")
            return

        row = self.db.execute(
            "SELECT * FROM customers WHERE id=?", (self.selected_id,)
        ).fetchone()

        win = tk.Toplevel(self)
        win.title("Edit Pelanggan")
        win.geometry("620x320")

        frame = tk.Frame(win)
        frame.pack(padx=20, pady=20, fill="both", expand=True)

        # ========================
        # LEFT SIDE
        # ========================

        tk.Label(frame, text="Nama").grid(row=0, column=0, sticky="w")
        name_entry = tk.Entry(frame, width=30)
        name_entry.insert(0, row["name"] or "")
        name_entry.grid(row=0, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Telepon").grid(row=1, column=0, sticky="w")
        phone_entry = tk.Entry(frame, width=30)
        phone_entry.insert(0, row["phone"] or "")
        phone_entry.grid(row=1, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="Email").grid(row=2, column=0, sticky="w")
        email_entry = tk.Entry(frame, width=30)
        email_entry.insert(0, row["email"] or "")
        email_entry.grid(row=2, column=1, pady=5, padx=(0, 20))

        tk.Label(frame, text="NPWP").grid(row=3, column=0, sticky="w")
        npwp_entry = tk.Entry(frame, width=30)
        npwp_entry.insert(0, row["npwp"] or "")
        npwp_entry.grid(row=3, column=1, pady=5, padx=(0, 20))

        # ========================
        # RIGHT SIDE
        # ========================

        tk.Label(frame, text="Limit Kredit").grid(row=0, column=2, sticky="w")
        credit_entry = tk.Entry(frame, width=30)
        credit_entry.insert(0, row["credit_limit"] or 0)
        credit_entry.grid(row=0, column=3, pady=5)

        bind_number_entry(credit_entry)

        tk.Label(frame, text="Sales PIC").grid(row=1, column=2, sticky="w")
        sales_entry = tk.Entry(frame, width=30)
        sales_entry.insert(0, row["sales_pic"] or "")
        sales_entry.grid(row=1, column=3, pady=5)

        tk.Label(frame, text="Domisili").grid(row=2, column=2, sticky="w")
        dom_entry = tk.Entry(frame, width=30)
        dom_entry.insert(0, row["domisili"] or "")
        dom_entry.grid(row=2, column=3, pady=5)

        tk.Label(frame, text="Alamat").grid(row=3, column=2, sticky="w")
        addr_entry = tk.Entry(frame, width=30)
        addr_entry.insert(0, row["address"] or "")
        addr_entry.grid(row=3, column=3, pady=5)

        inactive_var = tk.IntVar(value=0 if row["is_active"] == 1 else 1)

        tk.Checkbutton(frame, text="Nonaktifkan Pelanggan", variable=inactive_var).grid(
            row=4, column=3, sticky="w"
        )

        # ========================
        # SAVE
        # ========================

        def save():

            name = name_entry.get().strip()
            phone = phone_entry.get().strip()
            email = email_entry.get().strip()
            npwp = npwp_entry.get().strip()

            credit = parse_number(credit_entry.get())
            sales = sales_entry.get().strip()
            dom = dom_entry.get().strip()
            addr = addr_entry.get().strip()

            if not name:
                messagebox.showerror("Error", "Nama pelanggan wajib diisi")
                return

            self.db.execute(
                """
            UPDATE customers
            SET name=?,phone=?,email=?,npwp=?,credit_limit=?,
                sales_pic=?,domisili=?,address=?,is_active=?
            WHERE id=?
            """,
                (
                    name,
                    phone,
                    email,
                    npwp,
                    credit,
                    sales,
                    dom,
                    addr,
                    0 if inactive_var.get() else 1,
                    self.selected_id,
                ),
                commit=True,
            )

            messagebox.showinfo("Sukses", "Pelanggan berhasil diupdate")

            win.destroy()
            self.load_data()

        tk.Button(frame, text="Update", bg="#06b6d4", width=14, command=save).grid(
            row=5, column=3, pady=20
        )

    # ======================================
    # DELETE
    # ======================================

    def delete_customer(self):

        # jika ada checkbox dipilih
        if self.selected_customers:
            targets = list(self.selected_customers)

        else:
            # jika tidak ada checkbox gunakan baris yang dipilih
            sel = self.tree.selection()

            if not sel:
                messagebox.showerror("Error", "Pilih atau centang pelanggan dulu")
                return

            vals = self.tree.item(sel[0], "values")
            targets = [vals[2]]  # kode pelanggan

        win = tk.Toplevel(self)
        win.title("Konfirmasi")
        win.geometry("350x150")
        win.configure(bg="#e5e7eb")

        tk.Label(
            win,
            text="Hapus Pelanggan?",
            font=("Segoe UI", 14, "bold"),
            bg="#e5e7eb",
        ).pack(pady=20)

        btn_frame = tk.Frame(win, bg="#e5e7eb")
        btn_frame.pack()

        def confirm_delete():

            for code in targets:

                row = self.db.execute(
                    "SELECT id FROM customers WHERE code=?", (code,)
                ).fetchone()

                if row:
                    self.db.execute(
                        "DELETE FROM customers WHERE id=?", (row["id"],), commit=True
                    )

            self.selected_customers.clear()

            win.destroy()

            messagebox.showinfo("Sukses", "Pelanggan berhasil dihapus")

            self.load_data()

        tk.Button(
            btn_frame,
            text="Ya",
            width=10,
            bg="#16a34a",
            fg="white",
            command=confirm_delete,
        ).pack(side="left", padx=10)

        tk.Button(
            btn_frame,
            text="Tidak",
            width=10,
            bg="#dc2626",
            fg="white",
            command=win.destroy,
        ).pack(side="left", padx=10)

    def refresh_checkboxes(self):

        for item in self.tree.get_children():

            vals = list(self.tree.item(item, "values"))
            code = vals[2]

            if code in self.selected_customers:
                vals[0] = "☑"
            else:
                vals[0] = "☐"

            self.tree.item(item, values=vals)

    def toggle_checkbox(self, event):

        region = self.tree.identify("region", event.x, event.y)

        if region != "cell":
            return

        column = self.tree.identify_column(event.x)

        if column != "#1":
            return

        item = self.tree.identify_row(event.y)

        if not item:
            return

        vals = list(self.tree.item(item, "values"))
        code = vals[2]

        if code in self.selected_customers:
            self.selected_customers.remove(code)
        else:
            self.selected_customers.add(code)

        self.refresh_checkboxes()

    def print_customers(self):

        win = tk.Toplevel(self)
        win.title("Konfirmasi")
        win.geometry("350x150")
        win.configure(bg="#e5e7eb")

        tk.Label(
            win,
            text="Cetak Data Pelanggan?",
            font=("Segoe UI", 14, "bold"),
            bg="#e5e7eb",
        ).pack(pady=20)

        btn_frame = tk.Frame(win, bg="#e5e7eb")
        btn_frame.pack()

        def confirm_print():

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")],
                title="Simpan Data Pelanggan",
            )

            if not filepath:
                return

            rows = self.db.execute(
                """
                SELECT code,name,phone,email,npwp,
                    sales_pic,domisili,address,credit_limit,is_active
                FROM customers
                ORDER BY name
                """
            ).fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = "Pelanggan"

            headers = [
                "Kode",
                "Nama",
                "Telepon",
                "Email",
                "NPWP",
                "Sales PIC",
                "Domisili",
                "Alamat",
                "Limit Kredit",
                "Status",
            ]

            ws.append(headers)

            for r in rows:

                status = "AKTIF" if r["is_active"] == 1 else "NONAKTIF"

                ws.append(
                    [
                        r["code"],
                        r["name"],
                        r["phone"],
                        r["email"],
                        r["npwp"],
                        r["sales_pic"],
                        r["domisili"],
                        r["address"],
                        r["credit_limit"],
                        status,
                    ]
                )

            wb.save(filepath)

            win.destroy()

            messagebox.showinfo("Sukses", "Data pelanggan berhasil dicetak")

        tk.Button(
            btn_frame,
            text="Ya",
            width=10,
            bg="#16a34a",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=confirm_print,
        ).pack(side="left", padx=10)

        tk.Button(
            btn_frame,
            text="Tidak",
            width=10,
            bg="#dc2626",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            command=win.destroy,
        ).pack(side="left", padx=10)

    def import_customers(self):

        filepath = filedialog.askopenfilename(
            filetypes=[("Excel File", "*.xlsx")], title="Import Data Pelanggan"
        )

        if not filepath:
            return

        wb = load_workbook(filepath)
        ws = wb.active

        insert_count = 0
        update_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):

            code = row[0]
            name = row[1]
            phone = row[2]
            email = row[3]
            npwp = row[4]
            sales = row[5]
            dom = row[6]
            addr = row[7]
            credit = row[8] or 0
            status = row[9] if len(row) > 9 else "AKTIF"

            is_active = 1
            if str(status).upper() == "NONAKTIF":
                is_active = 0

            if not code or not name:
                continue

            existing = self.db.execute(
                "SELECT id FROM customers WHERE code=?", (code,)
            ).fetchone()

            if existing:

                self.db.execute(
                    """
                UPDATE customers
                SET name=?,phone=?,email=?,npwp=?,sales_pic=?,
                    domisili=?,address=?,credit_limit=?,is_active=?
                WHERE code=?
                """,
                    (
                        name,
                        phone,
                        email,
                        npwp,
                        sales,
                        dom,
                        addr,
                        credit,
                        is_active,
                        code,
                    ),
                    commit=True,
                )

                update_count += 1

            else:

                self.db.execute(
                    """
                INSERT INTO customers
                (code,name,phone,email,npwp,sales_pic,domisili,address,credit_limit,is_active)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                """,
                    (
                        code,
                        name,
                        phone,
                        email,
                        npwp,
                        sales,
                        dom,
                        addr,
                        credit,
                        is_active,
                    ),
                    commit=True,
                )

                insert_count += 1

        messagebox.showinfo(
            "Import selesai",
            f"Pelanggan baru : {insert_count}\nUpdate : {update_count}",
        )

        self.load_data()
